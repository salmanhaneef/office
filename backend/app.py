from flask import Flask, render_template, request, redirect, url_for, session, jsonify
from pymongo import MongoClient
import bcrypt
from flask_cors import CORS
from datetime import timedelta
import pdfplumber
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Alignment, Font
from collections import namedtuple
import re
import os 
from flask import Flask, send_file
# other imports

app = Flask(__name__)

# Enable CORS and allow credentials to be shared with frontend


# Enable CORS with credentials support
CORS(app, supports_credentials=True, origins=["http://localhost:5173"])


# Secret key for session management
app.secret_key = "Company1"

# Configure session cookie settings
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'  # Use 'Strict' if stricter security is needed
app.config['SESSION_COOKIE_SECURE'] = False  # Set to True in production with HTTPS
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(minutes=30)  # Session lifetime

# MongoDB connection URI
client = MongoClient("mongodb+srv://salman:12SAlm@cluster0.lq7um.mongodb.net/company1?retryWrites=true&w=majority&appName=Cluster0")
# client = MongoClient("mongodb://localhost:27017/project")
db = client.get_database('total_records')
records = db.register
Line = namedtuple('Line', 'Commercial_Package_Policy Premium_Policy Premium_Policy1')
CoverageDetails = namedtuple('CoverageDetails', 'Coverage Premium1 Premium2 Premium3 Premium1_1 Premium2_1 Premium3_1')
Line1 = namedtuple('Line1', 'Coverage_Type Limit Premium Limit1 Premium1')
PremiumDetails = namedtuple('PremiumDetails', 'Premium_Coverage Premium1 Premium2 Premium3 Premium1_1 Premium2_1 Premium3_1')

UPLOAD_FOLDER = 'uploads'  
ALLOWED_EXTENSIONS = {'pdf'}  
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER  

# Create the upload directory if it does not exist  
if not os.path.exists(UPLOAD_FOLDER):  
    os.makedirs(UPLOAD_FOLDER)  

def allowed_file(filename):  
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


@app.route('/signup', methods=['POST'])  
def signup():  
    # Use request.get_json() to handle JSON content  
    data = request.get_json()  

    user = data.get("fullname")  
    email = data.get("email")  
    password1 = data.get("password1")  
    password2 = data.get("password2")  

    # Basic validation  
    if not user or not email or not password1 or not password2:  
        return jsonify({"message": "All fields are required!"}), 400  

    # Check if the user or email already exists  
    user_found = records.find_one({"name": user})  
    email_found = records.find_one({"email": email})  

    if user_found:  
        return jsonify({"message": "There already is a user by that name"}), 400  
    if email_found:  
        return jsonify({"message": "This email already exists in the database"}), 400  
    if password1 != password2:  
        return jsonify({"message": "Passwords should match!"}), 400  

    # Hash the password  
    hashed = bcrypt.hashpw(password1.encode('utf-8'), bcrypt.gensalt())  

    # Insert user into MongoDB  
    user_input = {'name': user, 'email': email, 'password': hashed}  
    
    try:  
        records.insert_one(user_input)  
    except Exception as e:  
        return jsonify({"message": "An error occurred while registering the user.", "error": str(e)}), 500  

    # Set session for the user  
    session["email"] = email  
    session.permanent = True  

    return jsonify({"success": True, "message": "User registered successfully!"}), 200
@app.route("/login", methods=["POST"])
def login():
    # If the user is already logged in, redirect to the logged_in route
    if "email" in session:
        print("Already logged in as:", session["email"])
        return redirect(url_for("logged_in"))

    if request.method == "POST":
        email = request.json.get("email")
        password = request.json.get("password")

        # Basic validation
        if not email or not password:
            return jsonify({"message": "Email and password are required!"}), 400

        # Check if the email exists
        email_found = records.find_one({"email": email})

        if email_found:
            passwordcheck = email_found['password']

            # Check if the password matches
            if bcrypt.checkpw(password.encode('utf-8'), passwordcheck):
                session["email"] = email
                session.permanent = True  # Make the session permanent
                print("Login successful, session set for:", email)
                return jsonify({"success": True}), 200
            else:
                return jsonify({"message": "Wrong password"}), 400
        else:
            return jsonify({"message": "Email not found"}), 400

    return jsonify({"message": "Invalid request method"}), 405

@app.route('/logged_in')
def logged_in():
    # Check if the session is active
    if "email" in session:
        print("Logged in as:", session["email"])
        return jsonify({"email": session["email"]}), 200
    else:
        print("No session found")
        return jsonify({"message": "Unauthorized"}), 401

@app.route("/logout", methods=["POST", "GET"])
def logout():
    print("Logout request received.")
    print("Session contents:", session)
    if "email" in session:
        print("Session exists for:", session["email"], "logging out...")
        session.pop("email", None)  # Remove the session
        return jsonify({"success": True}), 200
    else:
        print("No active session found.")
        return jsonify({"message": "No active session"}), 400
@app.route('/upload', methods=['POST'])
def upload_pdf():
    if 'file' not in request.files:
        return jsonify({"error": "No file part"}), 400
    
    file = request.files['file']
    
    if file.filename == '':
        return jsonify({"error": "No selected file"}), 400
    
    if file and file.filename.endswith('.pdf'):
        file_path = os.path.join('uploads', file.filename)
        file.save(file_path)
        
        excel_file = process_pdf(file_path)
        
        os.remove(file_path)
        
        return jsonify({"message": "PDF processed successfully", "file": excel_file}), 200
    
    return jsonify({"error": "Invalid file type"}), 400
def process_pdf(file):
    # Initialize named tuples and regex patterns
    ###complete code
 

    # Define named tuples to hold the extracted data  
    Person = namedtuple('Person', 'Name DOB Status')  
    Vehicle = namedtuple('Vehicle', 'Description VIN')
    CoverageDetails = namedtuple('CoverageDetails', 'Coverage Premium1 Premium2 Premium3')
    Line = namedtuple('Line', 'Commercial_Package_Policy Premium')
    Premises = namedtuple('Premises', 'Premises_Number Address Blanket_and_Limit Valuation Coinsurance Inflation_Guard')
    Coverages = namedtuple('Coverages', 'Coverage LIMITS Notes')
    Line1 = namedtuple('Line1', 'Coverage_Type Limit Premium')
    

    # Regex patterns
    premises_address_re = re.compile(r'(\d{3})\s+(.*?(?:\s+PA,\s+\d{5}))')
    premises_details_re = re.compile(r'(\d{3})\s+\$?([\d,]+)\s+([\w\s-]*)\s+(\d+)%\s+(\d+)%\s*([\w\s-]*)?')
    coverage_re = re.compile(  
        r'([A-Za-z\s\–]+(?:\n[A-Za-z\s\–]+)*)\s+\$(\d{1,3}(?:,\d{3})*)(?:\s+\w+)?'  
        r'([A-Za-z\s]+(?:\n[A-Za-z\s]+)*)?$'  # Capture any additional notes or terms  
    )  

    pattern68 = re.compile(
        r'([A-Za-z\s\–/]+)\s+\$(\d{1,3}(?:,\d{3})*)\s+per Occurrence\s+\$\s*(\d{1,3}(?:,\d{3})*)\s+per Occurrence'
    )
    
    person_re = re.compile(r'([A-Z\s]+)\s+(\d{2}/\d{2}/\d{4})\s+(Insured on This Policy)')  
    vehicle_re = re.compile(r'(\d{4}\s[A-Z\s]+\s[A-Z\s]+)\s([0-9A-Z]{17})')
    amount_re = re.compile(r'([A-Za-z\s,]+?)(?:\s+\$([\d,]+\.\d{2})| Not Covered|:\s*\$.00)')  
    total_re = re.compile(r'Estimated Total Premium:\s*\$([\d,]+\.\d{2})')
    amount_re1 = re.compile(r'([A-Za-z\s/()]+)\s+(\$\d{1,3}(?:,\d{3})*(?:/\s*\$?\d{1,3}(?:,\d{3})*)?)?\s+(\$\d{1,3}(?:,\d{3})*(?:/\s*\$?\d{1,3}(?:,\d{3})*)?)?\s+(\$\d{1,3}(?:,\d{3})*(?:/\s*\$?\d{1,3}(?:,\d{3})*)?)?(?:\s+Included)?')  
    total_re1 = re.compile(r'([A-Za-z\s,]+?)(?:\s+\$([\d,]+\.\d{2})(?:\s+\$([\d,]+\.\d{2}))(?:\s+\$([\d,]+\.\d{2})))')
    ###
    pattern = re.compile(r'^(.*?)\s+\$(.*?)\s+(\d+(?:,\d+)?(?:\.\d+)?)\s*(\d+|$)', re.MULTILINE)  
    vehicle_premium_pattern = re.compile(r"Total premium for \d{4} [A-Z]+\s*[A-Z]*\s*\$(\d+(?:,\d{3})*(?:\.\d{2})?)")  
    total_policy_premium_pattern = re.compile(r"Total Policy Premium:\s*\$(\d+(?:,\d{3})*(?:\.\d{2})?)")  
    subtotal_policy_premium_pattern = re.compile(r"Subtotal policy premium\s*\$(\d+(?:,\d{3})*(?:\.\d{2})?)")  
    total_6_month_premium_pattern = re.compile(r"Total 6 month policy premium and fees\s*\$(\d+(?:,\d{3})*(?:\.\d{2})?)") 
    premium_pattern = re.compile(r'Premium by Vehicle\s+(\$[\d,]+\.\d{2})\s+(\$[\d,]+\.\d{2})\s+(\$[\d,]+\.\d{2})')
    # Initialize lists to hold the extracted data  
    persons_list = []  
    vehicles_list = []  
    lines_list = []
    lines_list1 = []
    lines_list2 = []
    premises_list = []
    coverages_list = []
    Named_Insured_Mailing_Address = []  
    Policy_Number = []  
    Effective = []
    premiums = []  
    found_named_insured = False  
    found_policy_number = False  
    found_effective_from = False

    # Extract text from the PDF  
    with pdfplumber.open(file) as pdf:  
        for page_number, page in enumerate(pdf.pages, start=1):  
            text = page.extract_text()  
            if text:  
                lines = text.split('\n')
                ##
                index1 = [i + 1 for i, x in enumerate(lines) if "LOSS SUSTAINED CRIME COVERAGE PART DECLARATIONS" in x]
                if index1:
                    coverages_list.append(Coverages(Coverage="LOSS SUSTAINED CRIME COVERAGE PART DECLARATIONS", LIMITS=None, Notes=None))
                    start_line = index1[0]
                    for line in lines[start_line:]:
                        match3=pattern68.match(line)
                        if match3:
                            coverages_list.append(Coverages(Coverage=match3.group(1).strip(), LIMITS=f"${match3.group(2)} per Occurrence", Notes=None))
                
                index = [i + 1 for i, x in enumerate(lines) if "RELIGIOUS ORGANIZATION MANAGEMENT LIABILITY" in x]
                if index:
                    pre_index =[i + 1 for i, x in enumerate(lines) if "COVERAGE PART DECLARATIONS" in x]
                    start_line = index[0]
                    if pre_index:
                        coverages_list.append(Coverages(Coverage="RELIGIOUS ORGANIZATION MANAGEMENT LIABILITY COVERAGE PART DECLARATIONS", LIMITS=None, Notes=None))
                        start_line = index[0]
                        for line in lines[start_line:]:
                            match = coverage_re.match(line)
                            if match:
                                coverage_title = match.group(1).strip()
                                limits = match.group(2).strip()
                                notes = match.group(3).strip() if match.group(3) is not None else ""
                                coverages_list.append(Coverages(Coverage=coverage_title, LIMITS=limits, Notes=notes))
                general_liability_index = [i + 1 for i, x in enumerate(lines) if "GENERAL LIABILITY COVERAGE PART DECLARATIONS" in x] 
                if general_liability_index:
                    
                    coverages_list.append(Coverages(Coverage="GENERAL LIABILITY COVERAGE PART DECLARATIONS", LIMITS=None, Notes=None))
                    start_line = general_liability_index[0]
                    for line in lines[start_line:]:
                        match = coverage_re.match(line)
                        if match:
                            coverage_title = match.group(1).strip()
                            limits = match.group(2).strip()
                            notes = match.group(3).strip() if match.group(3) is not None else ""
                            coverages_list.append(Coverages(Coverage=coverage_title, LIMITS=limits, Notes=notes))
                violent_event_index = [i + 1 for i, x in enumerate(lines) if "VIOLENT EVENT EXPENSE COVERAGE" in x]
                if violent_event_index:
                    coverages_list.append(Coverages(Coverage="VIOLENT EVENT EXPENSE COVERAGE", LIMITS=None, Notes=None))
                    start_line = violent_event_index[0]
                    for line in lines[start_line:]:
                        match = coverage_re.match(line)
                        if match:
                            coverage_title = match.group(1).strip()
                            limits = match.group(2).strip()
                            notes = match.group(3).strip() if match.group(3) is not None else ""
                            coverages_list.append(Coverages(Coverage=coverage_title, LIMITS=limits, Notes=notes))
                religious_counseling_index = [i + 1 for i, x in enumerate(lines) if "RELIGIOUS COUNSELING SERVICES LIABILITY" in x]
                if religious_counseling_index:
                    start_line = religious_counseling_index[0]
                    religious_counseling_index1 = [i + 1 for i, x in enumerate(lines) if "COVERAGE" in x]
                    if religious_counseling_index1:
                        coverages_list.append(Coverages(Coverage="RELIGIOUS COUNSELING SERVICES LIABILITY COVERAGE", LIMITS=None, Notes=None))
                        start_line = religious_counseling_index1[0]
                        for line in lines[start_line:]:
                            match = coverage_re.match(line)
                            if match:
                                coverage_title = match.group(1).strip()
                                limits = match.group(2).strip()
                                notes = match.group(3).strip() if match.group(3) is not None else ""
                                coverages_list.append(Coverages(Coverage=coverage_title, LIMITS=limits, Notes=notes))
                index1 = [i + 1 for i, x in enumerate(lines) if "Primary use of the vehicle: Pleasure/Personal" in x]
                if index1:
                    start_line = index1[0]
                    for line in lines[start_line:]:
                        match3 = pattern.match(line)  
                        vehicle_premium_match = vehicle_premium_pattern.search(line)  
                        total_policy_premium_match = total_policy_premium_pattern.search(line)  
                        subtotal_policy_premium_match = subtotal_policy_premium_pattern.search(line)  
                        total_6_month_premium_match = total_6_month_premium_pattern.search(line)
                        # Capture coverage lines  
                        if match3:  
                            coverage_title = match3.group(1).strip()  
                            limits = match3.group(2).strip()  
                            notes = match3.group(3).strip() if match3.group(3) is not None else ""  
                            lines_list2.append(Line1(Coverage_Type=coverage_title, Limit=limits, Premium=notes))  

                        # Check for vehicle premium match and handle accordingly  
                        if vehicle_premium_match:  
                            coverage_title = "Vehicle Premium"   
                            limits = None  
                            notes = vehicle_premium_match.group(1).strip() if vehicle_premium_match else ""  
                            lines_list2.append(Line1(Coverage_Type=coverage_title, Limit=limits, Premium=notes))  
                        
                        # Capture total policy premium  
                        # Capture subtotal policy premium  
                        if subtotal_policy_premium_match:  
                            coverage_title = "Subtotal Policy Premium"   
                            limits = None  
                            notes = subtotal_policy_premium_match.group(1).strip() if subtotal_policy_premium_match else ""  
                            lines_list2.append(Line1(Coverage_Type=coverage_title, Limit=limits, Premium=notes))  

                        # Capture total 6 month premium  
                        if total_6_month_premium_match:  
                            coverage_title = "Total 6 Month Premium"   
                            limits = None  
                            notes = total_6_month_premium_match.group(1).strip() if total_6_month_premium_match else ""  
                            lines_list2.append(Line1(Coverage_Type=coverage_title, Limit=limits, Premium=notes))
                        
                for line in lines:
                    match_address = premises_address_re.search(line)
                    if match_address:
                        premises_number = match_address.group(1).strip()
                        premises_address = match_address.group(2).strip()
                        premises_list.append(Premises(premises_number, premises_address, '', '', '', ''))

                # Process premises details
                merged_lines = []
                temp_line = ''
                for line in lines:
                    if re.match(r'^\d{3}\s+', line):
                        if temp_line:
                            merged_lines.append(temp_line.strip())
                        temp_line = line
                    else:
                        temp_line += ' ' + line
                if temp_line:
                    merged_lines.append(temp_line.strip())

                data = "\n".join(merged_lines)
                matches = premises_details_re.findall(data)

                for match in matches:
                    premises_number = match[0].strip()
                    blanket_and_limit = f"${match[1].strip()}"
                    valuation = match[2].strip()
                    coinsurance = match[3] + '%'
                    inflation_guard = match[4] + '%'
                    premises_list.append(Premises(premises_number, '', blanket_and_limit, valuation, coinsurance, inflation_guard))

                
                # Extract Named Insured and Mailing Address
                pre_index = [i + 1 for i, x in enumerate(lines) if "Named Insured and Mailing Address:" in x]  
                pre_index1 = [i + 1 for i, x in enumerate(lines) if "Named Insured:" in x]  
                pre_index2 = [i + 1 for i, x in enumerate(lines) if "Drivers and household residents" in x]
                if pre_index:  
                    Named_Insured_Mailing_Address.append(lines[pre_index[0]])  
                    found_named_insured = True
                    # if found_named_insured:
                    #     break
                elif pre_index1:  
                    Named_Insured_Mailing_Address.append(lines[pre_index1[0]])  
                    found_named_insured = True  
                elif pre_index2:  
                    Named_Insured_Mailing_Address.append(lines[pre_index2[0]])  
                    found_named_insured = True

                # Extract Policy Number
                if not found_policy_number:  
                    pro_index = [i for i, x in enumerate(lines) if "Policy Number:" in x]
                    pro1_index = [i+1 for i, x in enumerate(lines) if "Policy Number:" in x]
                    if pro_index:
                        if lines[pro_index[0]] != "Policy Number:":
                            Policy_Number.append(lines[pro_index[0]])
                            found_policy_number = True        
                        else:
                            Policy_Number.append(lines[pro1_index[0]])  
                            found_policy_number = True
                
                # Extract Effective Dates
                if not found_effective_from:  
                    pr_index = [i for i, x in enumerate(lines) if "Policy Effective " in x]  
                    pr_index1 = [i for i, x in enumerate(lines) if "Policy Period:" in x]  
                    if pr_index:  
                        Effective.append(lines[pr_index[0]])  
                        if pr_index[0] + 1 < len(lines):  
                            Effective.append(lines[pr_index[0] + 1])  
                        found_effective_from = True  
                    if pr_index1:  
                        Effective.append(lines[pr_index1[0]])  
                        if pr_index1[0] + 1 < len(lines):  
                            Effective.append(lines[pr_index1[0] + 1])  
                        found_effective_from = True
                # if all([found_named_insured, found_policy_number, found_effective_from]):
                #     break 
                

                # Extract Coverage Details
                index = [i for i, x in enumerate(lines) if "This policy consists of the following coverage parts: " in x]
                if index:
                    start_line = index[0] + 1
                    for line in lines[start_line:]:
                        match2 = amount_re.search(line)
                        if match2:
                            coverage = match2.group(1).strip()
                            premium = match2.group(2).strip() if match2.group(2) else "Not Covered"
                            lines_list1.append(Line(Commercial_Package_Policy=coverage, Premium=premium))
                        
                        match3 = total_re.search(line)
                        if match3:
                            total_premium = match3.group(1).strip()
                            lines_list1.append(Line("Estimated Total Premium", total_premium))
                
                
                for line in lines:
                    match1 = total_re1.search(line)
                    match = amount_re1.search(line)

                    if match:  
                        coverage = match.group(1).strip()  
                        premium1 = match.group(2).strip() if match.group(2) else None  
                        premium2 = match.group(3).strip() if match.group(3) else None  
                        premium3 = match.group(4).strip() if match.group(4) else None  
                        lines_list.append(CoverageDetails(Coverage=coverage, Premium1=premium1, Premium2=premium2, Premium3=premium3))

                    if match1:  
                        coverage = match1.group(1).strip()  
                        premium1 = match1.group(2).strip() if match1.group(2) else None
                        premium2 = match1.group(3).strip() if match1.group(3) else None  
                        premium3 = match1.group(4).strip() if match1.group(4) else None  
                        lines_list.append(CoverageDetails(Coverage=coverage, Premium1=premium1, Premium2=premium2, Premium3=premium3))
                premium_match = premium_pattern.search(text)  
                if premium_match:  
                    premiums.append({'Premium1': premium_match.group(1), 'Premium2': premium_match.group(2), 'Premium3': premium_match.group(3)})

                # Extract Persons and Vehicles Information
                for line in lines:
                    person_match = person_re.search(line)  
                    vehicle_match = vehicle_re.search(line)

                    if person_match:  
                        name, dob, status = person_match.groups()  
                        persons_list.append(Person(name.strip(), dob.strip(), status.strip()))  

                    if vehicle_match:  
                        description, vin = vehicle_match.groups()  
                        vehicles_list.append(Vehicle(description.strip(), vin.strip())) 

    # Convert lists to DataFrames
    premium_df = pd.DataFrame(premiums)
    persons_df = pd.DataFrame(persons_list)
    vehicles_df = pd.DataFrame(vehicles_list)
    coverage_df1 = pd.DataFrame(lines_list1)
    coverage_df0 = pd.DataFrame(lines_list)
    coverage_df = pd.DataFrame(coverages_list)
    df = pd.DataFrame(lines_list2)
    def convert_to_float(premium_str):  
        return float(premium_str.replace('$', '').replace(',', ''))  

    # Calculate total premium if the DataFrame is not empty  
    if not premium_df.empty:  
        premium_df['Total Premium'] = premium_df.apply(lambda row: sum([convert_to_float(row['Premium1']),  
                                                                        convert_to_float(row['Premium2']),  
                                                                        convert_to_float(row['Premium3'])]), axis=1)  
        
        # Add a new first column with the value "Premium by Vehicle:"  
        premium_df.insert(0, 'Description', 'Premium by Vehicle:')
    # Print the DataFrames
    # Determine the maximum length among the lists
    max_length = max(len(Named_Insured_Mailing_Address), len(Policy_Number), len(Effective))

    # Extend the lists to match the maximum length
    Named_Insured_Mailing_Address.extend([None] * (max_length - len(Named_Insured_Mailing_Address)))
    Policy_Number.extend([None] * (max_length - len(Policy_Number)))
    Effective.extend([None] * (max_length - len(Effective)))

    # Create the DataFrame
    policy_data = pd.DataFrame({
        'Named Insured': Named_Insured_Mailing_Address,
        "Policy Number": Policy_Number,
        "Effective": Effective,
    }).drop_duplicates()
    ##0
    print("\nPremium DataFrame:")  
    print(premium_df)
    ##-1
    print("\ncoveragesss DataFrame:")
    print(df)
    # Print the DataFrame for verification
    ##1dd
    print("Policy_data")
    print(policy_data)
    ##2
    print("Persons DataFrame:")
    print(persons_df)
    ##3
    print("\nVehicles DataFrame:")
    print(vehicles_df)
    ##4dd
    print("\nPolicy DataFrame:")
    print(coverage_df1)
    ##5
    print("\nPolicy6 DataFrame:")
    print(coverage_df0)
    ##6dd

    print("\nCoverages DataFrame:")
    print(coverage_df)
    ##7dd
    premises_df = pd.DataFrame(premises_list).drop_duplicates()
    print("Premises DataFrame:")
    print(premises_df)
    excel_file = 'insurance_report_combined.xlsx'  
    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:  
        policy_data.to_excel(writer, index=False, sheet_name='Insurance Report', startrow=2)  
        
        # Access the workbook and the worksheet  
        workbook = writer.book  
        worksheet = writer.sheets['Insurance Report']  

        # Add headings for the policy report  
        report_heading = "Insurance Report"  
        worksheet.merge_cells('A1:C1')  
        cell = worksheet.cell(row=1, column=1)  
        cell.value = report_heading  
        cell.alignment = Alignment(horizontal='center', vertical='center')  
        cell.font = Font(bold=True)  

        # Add border style for the columns  
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))  
        
        min_width = 20  # Minimum column width  

        # Set column widths and apply styles  
        for col in worksheet.columns:  
            if not col:  # If the column is empty, skip  
                continue  
            
            max_length = 0  
            column_letter = get_column_letter(col[0].column)  

            for cell in col:  
                if cell.value:  # Check if the cell is not empty  
                    max_length = max(max_length, len(str(cell.value)))  # Update max_length  
                
            adjusted_width = max(max_length + 2, min_width)  # Add some extra space  
            worksheet.column_dimensions[column_letter].width = adjusted_width  
            
            # Set borders for each cell in the column  
            for cell in col:  
                cell.border = border  

        # Make header row bold (now it’s row 3 after writing the policy data)  
        for cell in worksheet[2]:  # Row indexing starts from 1  
            cell.font = Font(bold=True)  # Make the header bold  

        # Loop to append additional DataFrames   
        last_row = worksheet.max_row + 2  
        for df in [premises_df, coverage_df, coverage_df1, coverage_df0, persons_df, premium_df, df, vehicles_df]:  
            df.to_excel(writer, index=False, sheet_name='Insurance Report', startrow=last_row) 
            
            

            # Update last_row after writing each DataFrame  
            last_row = worksheet.max_row + 2  

            # Apply the same styling and formatting to each appended DataFrame  
            for col in worksheet.iter_cols(min_row=last_row):  
                if not col:  # If the column is empty, skip  
                    continue  
                
                max_length = 0  
                column_letter = get_column_letter(col[0].column)
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    cell.border = border
        

                for cell in col:  
                    if cell.value:  # Only check non-empty cells  
                        max_length = max(max_length, len(str(cell.value)))  

                adjusted_width = max(max_length + 2, min_width)  
                worksheet.column_dimensions[column_letter].width = adjusted_width  

                # Apply borders  
                for cell in col:  
                    cell.border = border  

        # Save changes  
        workbook.save(excel_file)  

    print(f'Data successfully written to {excel_file}') 
 
 
    # (Omitted: Full PDF processing and Excel writing code)
    return excel_file

@app.route('/download/<filename>', methods=['GET'])
def download_file(filename):
    file_path = os.path.join( filename)
    return send_file(file_path, as_attachment=True)
# @app.route('/view/<filename>', methods=['GET'])
# def view_file(filename):
#     file_path = os.path.join( filename)
#     return send_file(file_path)
@app.route('/view/<filename>', methods=['GET'])  
def view_file(filename):  
    file_path = os.path.join(filename)  # Modify this as needed  
    if os.path.exists(file_path):  
        df = pd.read_excel(file_path)  
        html = df.to_html()  # Converts DataFrame to HTML  
        return html  # Render HTML directly  
    return "File not found", 404
@app.route('/upload1', methods=['POST'])  
def upload_files():  
    # Check if both files are provided in the request  
    if 'file1' not in request.files or 'file2' not in request.files:  
        return jsonify({"error": "Both PDF files are required"}), 400  

    file1 = request.files['file1']  
    file2 = request.files['file2']  

    # Check for empty filenames  
    if file1.filename == '' or file2.filename == '':  
        return jsonify({"error": "No selected file"}), 400  

    # Check if both files are allowed  
    if file1 and allowed_file(file1.filename) and file2 and allowed_file(file2.filename):  
        filename1 = secure_filename(file1.filename)  
        filename2 = secure_filename(file2.filename)  
        filepath1 = os.path.join(app.config['UPLOAD_FOLDER'], filename1)  
        filepath2 = os.path.join(app.config['UPLOAD_FOLDER'], filename2)  

        try:  
            # Save the uploaded files  
            file1.save(filepath1)  
            file2.save(filepath2)  

            # Call your processing function  
            excel_file = process_pdfs(filepath1, filepath2)  

            # Optionally remove the temporary uploaded files  
            os.remove(filepath1)  
            os.remove(filepath2)  

            return jsonify({"message": "PDFs processed successfully", "file": excel_file}), 200  
        except Exception as e:  
            # Handle any errors that occur during file processing  
            return jsonify({"error": "An error occurred while processing the files", "details": str(e)}), 500  

    return jsonify({"error": "Invalid file type"}), 400

def process_pdfs(file1, file2):
 


# Regex patterns for part 1
    amount_re = re.compile(r'([A-Za-z\s,]+?)(?:\s+\$([\d,]+\.\d{2})| Not Covered|:\s*\$.00)')
    total_re = re.compile(r'Estimated Total Premium:\s*\$([\d,]+\.\d{2})')
    amount_re2 = re.compile(r'([A-Za-z\s$$]+)\s+\$([\d,]+(?:\.\d{2})?)')
    amount_re1 = re.compile(r'([A-Za-z\s/()]+)\s+(\$\d{1,3}(?:,\d{3})*(?:/\s*\$?\d{1,3}(?:,\d{3})*)?)?\s+(\$\d{1,3}(?:,\d{3})*(?:/\s*\$?\d{1,3}(?:,\d{3})*)?)?\s+(\$\d{1,3}(?:,\d{3})*(?:/\s*\$?\d{1,3}(?:,\d{3})*)?)?(?:\s+Included)?')
    total_re1 = re.compile(r'([A-Za-z\s,]+?)(?:\s+\$([\d,]+\.\d{2})(?:\s+\$([\d,]+\.\d{2}))(?:\s+\$([\d,]+\.\d{2})))')
    premium_pattern = re.compile(r'Premium by Vehicle\s+(\$[\d,]+\.\d{2})\s+(\$[\d,]+\.\d{2})\s+(\$[\d,]+\.\d{2})')
    pattern = re.compile(r'^(.*?)\s+\$(.*?)\s+(\d+(?:,\d+)?(?:\.\d+)?)\s*(\d+|$)', re.MULTILINE)  
    vehicle_premium_pattern = re.compile(r"Total premium for \d{4} [A-Z]+\s*[A-Z]*\s*\$(\d+(?:,\d{3})*(?:\.\d{2})?)")  
    total_policy_premium_pattern = re.compile(r"Total Policy Premium:\s*\$(\d+(?:,\d{3})*(?:\.\d{2})?)")  
    subtotal_policy_premium_pattern = re.compile(r"Subtotal policy premium\s*\$(\d+(?:,\d{3})*(?:\.\d{2})?)")  
    total_6_month_premium_pattern = re.compile(r"Total 6 month policy premium and fees\s*\$(\d+(?:,\d{3})*(?:\.\d{2})?)")  

    # Initialize lists to hold the extracted data for part 1

    lines_list1 = []
    lines_list2 = []
    lines_list3 = []
    premiums_first_pdf = []  
    premiums_second_pdf = []
    lines_list_first_pdf = []  
    lines_list_second_pdf = []
    Named_Insured_Mailing_Address = []  
    Policy_Number = []  
    Effective = []
    found_named_insured = False  
    found_policy_number = False  
    found_effective_from = False
    # Process the first PDF file for part 1
    with pdfplumber.open(file1) as pdf:  
        for page in pdf.pages:  
            text = page.extract_text()  
            if text:
                
                premium_match = premium_pattern.search(text)  
                if premium_match:  
                    premium1 = premium_match.group(1)  
                    premium2 = premium_match.group(2)  
                    premium3 = premium_match.group(3)  
                    premiums_first_pdf.append((premium1, premium2, premium3))
                lines = text.split('\n')
                for line in lines:
                    match = amount_re1.search(line)
                    match1 = total_re1.search(line)
                    if match:
                        coverage = match.group(1).strip()  
                        premium1 = match.group(2).strip() if match.group(2) else None  
                        premium2 = match.group(3).strip() if match.group(3) else None  
                        premium3 = match.group(4).strip() if match.group(4) else None  
                        lines_list_first_pdf.append(CoverageDetails(Coverage=coverage, Premium1=premium1, Premium2=premium2, Premium3=premium3, Premium1_1=None, Premium2_1=None, Premium3_1=None))
                    elif match1:
                        coverage = match1.group(1).strip()
                        premium1 = match1.group(2).strip() if match1.group(2) else None
                        premium2 = match1.group(3).strip() if match1.group(3) else None
                        premium3 = match1.group(4).strip() if match1.group(4) else None
                        lines_list_first_pdf.append(CoverageDetails(Coverage=coverage, Premium1=premium1, Premium2=premium2, Premium3=premium3, Premium1_1=None, Premium2_1=None, Premium3_1=None))
                pre_index = [i + 1 for i, x in enumerate(lines) if "Named Insured and Mailing Address:" in x]  
                pre_index1 = [i + 1 for i, x in enumerate(lines) if "Named Insured:" in x]  
                pre_index2 = [i + 1 for i, x in enumerate(lines) if "Drivers and household residents" in x]
                if pre_index:  
                    Named_Insured_Mailing_Address.append(lines[pre_index[0]])  
                    found_named_insured = True
                    # if found_named_insured:
                    #     break
                elif pre_index1:  
                    Named_Insured_Mailing_Address.append(lines[pre_index1[0]])  
                    found_named_insured = True  
                elif pre_index2:  
                    Named_Insured_Mailing_Address.append(lines[pre_index2[0]])  
                    found_named_insured = True

                # Extract Policy Number
                if not found_policy_number:  
                    pro_index = [i for i, x in enumerate(lines) if "Policy Number:" in x]
                    pro1_index = [i+1 for i, x in enumerate(lines) if "Policy Number:" in x]
                    if pro_index:
                        if lines[pro_index[0]] != "Policy Number:":
                            Policy_Number.append(lines[pro_index[0]])
                            found_policy_number = True        
                        else:
                            Policy_Number.append(lines[pro1_index[0]])  
                            found_policy_number = True
                
                # Extract Effective Dates
                if not found_effective_from:  
                    pr_index = [i for i, x in enumerate(lines) if "Policy Effective " in x]  
                    pr_index1 = [i for i, x in enumerate(lines) if "Policy Period:" in x]  
                    if pr_index:  
                        Effective.append(lines[pr_index[0]])  
                        if pr_index[0] + 1 < len(lines):  
                            Effective.append(lines[pr_index[0] + 1])  
                        found_effective_from = True  
                    if pr_index1:  
                        Effective.append(lines[pr_index1[0]])  
                        if pr_index1[0] + 1 < len(lines):  
                            Effective.append(lines[pr_index1[0] + 1])  
                        found_effective_from = True
                #print("Extracted text for PDF 1:", lines)  # Debug print
                
                # Processing coverage parts
                index = [i for i, x in enumerate(lines) if "This policy consists of the following coverage parts: " in x]
                if index:
                    start_line = index[0] + 1
                    for line in lines[start_line:]:
                        match2 = amount_re.search(line)
                        if match2:
                            coverage = match2.group(1).strip()
                            premium = match2.group(2).strip() if match2.group(2) else "Not Covered"
                            lines_list1.append(Line(Commercial_Package_Policy=coverage, Premium_Policy=premium, Premium_Policy1=""))
                            #print(f"Matched coverage: {coverage}, premium: {premium}")  # Debug print

                        match3 = total_re.search(line)
                        if match3:
                            total_premium = match3.group(1).strip()
                            lines_list1.append(Line("Estimated Total Premium", total_premium, ""))
                index1 = [i + 1 for i, x in enumerate(lines) if "Primary use of the vehicle: Pleasure/Personal" in x]
                if index1:
                    start_line = index1[0]
                    for line in lines[start_line:]:
                        match4 = pattern.match(line)
                        vehicle_premium_match = vehicle_premium_pattern.search(line)
                        total_policy_premium_match = total_policy_premium_pattern.search(line)
                        subtotal_policy_premium_match = subtotal_policy_premium_pattern.search(line)
                        total_6_month_premium_match = total_6_month_premium_pattern.search(line)
                        
                        if match4:
                            coverage_title = match4.group(1).strip()
                            limits = match4.group(2).strip()
                            notes = match4.group(3).strip() if match4.group(3) is not None else ""
                            lines_list2.append(Line1(Coverage_Type=coverage_title, Limit=limits, Premium=notes, Limit1=None, Premium1=None))

                        if vehicle_premium_match:
                            lines_list2.append(Line1(Coverage_Type="Vehicle Premium", Limit=None, Premium=vehicle_premium_match.group(1).strip(), Limit1=None, Premium1=None))

                        if total_policy_premium_match:
                            lines_list2.append(Line1(Coverage_Type="Total Policy Premium", Limit=None, Premium=total_policy_premium_match.group(1).strip(), Limit1=None, Premium1=None))

                        if subtotal_policy_premium_match:
                            lines_list2.append(Line1(Coverage_Type="Subtotal Policy Premium", Limit=None, Premium=subtotal_policy_premium_match.group(1).strip(), Limit1=None, Premium1=None))

                        if total_6_month_premium_match:
                            lines_list2.append(Line1(Coverage_Type="Total 6 Month Premium", Limit=None, Premium=total_6_month_premium_match.group(1).strip(), Limit1=None, Premium1=None))
                            #print(f"Matched total premium: {total_premium}")  # Debug print
    df1 = pd.DataFrame(lines_list2)
            

    # Process the second PDF file and update Premium1 for part 1
    with pdfplumber.open(file2) as pdf:  
        for page in pdf.pages:  
            text = page.extract_text()  
            if text:
                premium_match = premium_pattern.search(text)  
                if premium_match:  
                    premium1_1 = premium_match.group(1)  
                    premium2_1 = premium_match.group(2)  
                    premium3_1 = premium_match.group(3)  
                    premiums_second_pdf.append((premium1_1, premium2_1, premium3_1))
                lines = text.split('\n')
                for line in lines:
                    match = amount_re1.search(line)
                    match1 = total_re1.search(line)
                    if match:
                        coverage = match.group(1).strip()  
                        premium1 = match.group(2).strip() if match.group(2) else None  
                        premium2 = match.group(3).strip() if match.group(3) else None  
                        premium3 = match.group(4).strip() if match.group(4) else None  
                        lines_list_second_pdf.append(CoverageDetails(Coverage=coverage, Premium1=premium1, Premium2=premium2, Premium3=premium3, Premium1_1=None, Premium2_1=None, Premium3_1=None))
                    elif match1:
                        coverage = match1.group(1).strip()
                        premium1 = match1.group(2).strip() if match1.group(2) else None
                        premium2 = match1.group(3).strip() if match1.group(3) else None
                        premium3 = match1.group(4).strip() if match1.group(4) else None
                        lines_list_second_pdf.append(CoverageDetails(Coverage=coverage, Premium1=premium1, Premium2=premium2, Premium3=premium3, Premium1_1=None, Premium2_1=None, Premium3_1=None))
                #print("Extracted text for PDF 2:", lines)  # Debug print
                

                # Processing coverage parts
                index = [i for i, x in enumerate(lines) if "This policy consists of the following coverage parts: " in x]
                if index:
                    start_line = index[0] + 1
                    for i, line in enumerate(lines[start_line:]):
                        match2 = amount_re.search(line)
                        if match2 and i < len(lines_list1):  
                            premium1 = match2.group(2).strip() if match2.group(2) else "Not Covered"
                            lines_list1[i] = lines_list1[i]._replace(Premium_Policy1=premium1)  
                            #print(f"Updated line {i} with Premium_Policy1: {premium1}")  # Debug print

                        match3 = total_re.search(line)
                        if match3 and i < len(lines_list1):  
                            total_premium1 = match3.group(1).strip()
                            if lines_list1[i].Commercial_Package_Policy == "Estimated Total Premium":
                                lines_list1[i] = lines_list1[i]._replace(Premium_Policy1=total_premium1)
                index2 = [i + 1 for i, x in enumerate(lines) if "Primary use of the vehicle: Pleasure/Personal" in x]
                if index2:
                    start_line = index2[0]
                    for line in lines[start_line:]:
                        match4 = pattern.match(line)
                        vehicle_premium_match = vehicle_premium_pattern.search(line)
                        total_policy_premium_match = total_policy_premium_pattern.search(line)
                        subtotal_policy_premium_match = subtotal_policy_premium_pattern.search(line)
                        total_6_month_premium_match = total_6_month_premium_pattern.search(line)
                        
                        if match3:
                            coverage_title = match4.group(1).strip()
                            limits = match4.group(2).strip()
                            notes = match4.group(3).strip() if match4.group(3) is not None else ""
                            lines_list3.append(Line1(Coverage_Type=coverage_title, Limit=limits, Premium=notes, Limit1=None, Premium1=None))

                        if vehicle_premium_match:
                            lines_list3.append(Line1(Coverage_Type="Vehicle Premium", Limit=None, Premium=vehicle_premium_match.group(1).strip(), Limit1=None, Premium1=None))

                        if total_policy_premium_match:
                            lines_list3.append(Line1(Coverage_Type="Total Policy Premium", Limit=None, Premium=total_policy_premium_match.group(1).strip(), Limit1=None, Premium1=None))

                        if subtotal_policy_premium_match:
                            lines_list3.append(Line1(Coverage_Type="Subtotal Policy Premium", Limit=None, Premium=subtotal_policy_premium_match.group(1).strip(), Limit1=None, Premium1=None))

                        if total_6_month_premium_match:
                            lines_list3.append(Line1(Coverage_Type="Total 6 Month Premium", Limit=None, Premium=total_6_month_premium_match.group(1).strip(), Limit1=None, Premium1=None))
                                #print(f"Updated total premium for Estimated Total Premium: {total_premium1}")  # Debug print


    df2 = pd.DataFrame(lines_list3)            
                
    # Convert to DataFrame for part 1
    coverage_df1 = pd.DataFrame(lines_list1)
    df = pd.DataFrame(lines_list2)
    # Verify DataFrame columns and data
    # print("DataFrame for part 1:", coverage_df1)

    # Check if DataFrame is empty
    if coverage_df1.empty:
        print("No data was collected for coverage_df1.")
    else:
        # Handle empty strings and non-numeric values in 'Premium' and 'Premium1' for part 1
        coverage_df1['Premium_Policy'] = coverage_df1['Premium_Policy'].replace(['Not Covered', ''], '0').str.replace(',', '').astype(float)
        coverage_df1['Premium_Policy1'] = coverage_df1['Premium_Policy1'].replace(['Not Covered', ''], '0').str.replace(',', '').astype(float)

        # Add a new column 'Difference' by subtracting Premium1 from Premium for part 1
        coverage_df1['Difference'] = coverage_df1['Premium_Policy'] - coverage_df1['Premium_Policy1']
    coverage_df2 = pd.DataFrame(lines_list_first_pdf)

    # Verify DataFrame columns for coverage_df2
    # print("DataFrame columns for part 2:", coverage_df2.columns)

    # Match and update the DataFrame with data from the second PDF for part 2
    for entry2 in lines_list_second_pdf:
        matched = False
        for i, row in coverage_df2.iterrows():
            if row['Coverage'] == entry2.Coverage:
                coverage_df2.at[i, 'Premium1_1'] = entry2.Premium1
                coverage_df2.at[i, 'Premium2_1'] = entry2.Premium2
                coverage_df2.at[i, 'Premium3_1'] = entry2.Premium3
                matched = True
                break
        if not matched:
            print(f"Coverage '{entry2.Coverage}' from second PDF did not match any entry in the first PDF.")

    # Add columns for differences for part 2
    coverage_df2['difference1_1'] = None
    coverage_df2['difference2_1'] = None
    coverage_df2['difference3_1'] = None

    # Calculate differences for part 2
    for i, row in coverage_df2.iterrows():
        def parse_premium(premium_str):
            if premium_str is None:
                return None
            numeric_parts = re.findall(r'\d+(?:,\d{3})*(?:\.\d{2})?', premium_str)
            if numeric_parts:
                return float(numeric_parts[0].replace(',', ''))
            return None

        premium1 = parse_premium(row['Premium1'])
        premium1_1 = parse_premium(row['Premium1_1'])
        premium2 = parse_premium(row['Premium2'])
        premium2_1 = parse_premium(row['Premium2_1'])
        premium3 = parse_premium(row['Premium3'])
        premium3_1 = parse_premium(row['Premium3_1'])
        
        coverage_df2.at[i, 'difference1_1'] = (premium1_1 - premium1) if premium1 is not None and premium1_1 is not None else None
        coverage_df2.at[i, 'difference2_1'] = (premium2_1 - premium2) if premium2 is not None and premium2_1 is not None else None
        coverage_df2.at[i, 'difference3_1'] = (premium3_1 - premium3) if premium3 is not None and premium3_1 is not None else None

    # Display the updated DataFrames

    print("Coverage DataFrame from part 1:")
    print(coverage_df1)
    print("\nCoverage DataFrame from part 2:")
    print(coverage_df2)

    min_length = min(len(premiums_first_pdf), len(premiums_second_pdf))  
    # print(f"Minimum length of premiums: {min_length}")  # Debug print  

    premiums_first_pdf = premiums_first_pdf[:min_length]  
    premiums_second_pdf = premiums_second_pdf[:min_length]  

    # Create PremiumDetails instances  
    premium_details_list = [  
        PremiumDetails(  
            Premium_Coverage="Premium of Vehicles",  
            Premium1=premiums_first_pdf[i][0],  
            Premium2=premiums_first_pdf[i][1],  
            Premium3=premiums_first_pdf[i][2],  
            Premium1_1=premiums_second_pdf[i][0],  
            Premium2_1=premiums_second_pdf[i][1],  
            Premium3_1=premiums_second_pdf[i][2]  
        )  
        for i in range(min_length)  
    ]  

    # Check if the premium details are populated  
    # print("Premium details list:", premium_details_list)  

    # Create a DataFrame from the list of PremiumDetails  
    premium_df = pd.DataFrame(premium_details_list)  

    # Verify the DataFrame structure  
    # print("DataFrame structure:\n", premium_df.head())  

    # Convert premium columns to numeric values, if they exist  
    premium_columns = ['Premium1', 'Premium2', 'Premium3', 'Premium1_1', 'Premium2_1', 'Premium3_1']  
    for col in premium_columns:  
        if col in premium_df.columns:  # Check if the column exists  
            premium_df[col] = premium_df[col].replace({r'\$': '', r',': ''}, regex=True).astype(float)  
        else:  
            print(f"Warning: Column {col} not found in the DataFrame.")  

    # Proceed if DataFrame has required columns  
    if len(premium_df.columns) > 0:  
        # Add TotalPremium1 and TotalPremium2 columns  
        premium_df['TotalPremium1'] = premium_df[['Premium1', 'Premium2', 'Premium3']].sum(axis=1)  
        premium_df['TotalPremium2'] = premium_df[['Premium1_1', 'Premium2_1', 'Premium3_1']].sum(axis=1)  

        # Add Difference column  
        premium_df['Difference'] = premium_df['TotalPremium1'] - premium_df['TotalPremium2']  

        # Print the final DataFrame  
        print(premium_df)  
    else:  
        print("No valid premium data found. DataFrame is empty.")   
    print("Columns in df1:", df1.columns.tolist())  
    print("Columns in df2:", df2.columns.tolist())
    print("DF1 contents:\n", df1.head())
    print("DF2 contents:\n", df2.head())

    # Check if 'Coverage_Type' is in both dataframes
    if 'Coverage_Type' not in df1.columns:
        print("Coverage_Type column not found in df1")

    if 'Coverage_Type' not in df2.columns:
        print("Coverage_Type column not found in df2")

    # Merge data from the two PDFs
    merged_df = df1.copy()

    # Update 'Limit1' and 'Premium1' columns from df2
    for index, row in merged_df.iterrows():
        corresponding_row = df2[df2['Coverage_Type'] == row['Coverage_Type']]
        if not corresponding_row.empty:
            merged_df.at[index, 'Limit1'] = corresponding_row['Limit'].values[0]
            merged_df.at[index, 'Premium1'] = corresponding_row['Premium'].values[0]

    # Check merged DataFrame structure
    print("Merged DataFrame structure before cleaning:\n", merged_df.head())
    print("Merged DataFrame columns:\n", merged_df.columns.tolist()) 

    # Function to clean and convert currency strings to floats
    def clean_currency(value):
        if isinstance(value, str):
            value = value.replace(',', '').replace('$', '')
            try:
                return float(value)
            except ValueError:
                return None
        return float(value) if value else None

    # Apply the cleaning function to Premium and Premium1 columns
    if 'Premium' in merged_df.columns and 'Premium1' in merged_df.columns:
        merged_df['Premium'] = merged_df['Premium'].apply(clean_currency)
        merged_df['Premium1'] = merged_df['Premium1'].apply(clean_currency)
        conditions = merged_df['Coverage_Type'].isin(["Vehicle Premium", "Subtotal Policy Premium", "Total 6 Month Premium"])
        merged_df.loc[conditions, 'Difference'] = merged_df['Premium1'] - merged_df['Premium']
    else:
        print("Necessary columns are missing in the merged DataFrame.")
    # max_length = max(len(Named_Insured_Mailing_Address), len(Policy_Number), len(Effective))

    # Extend the lists to match the maximum length
    Named_Insured_Mailing_Address.extend([None] * (max_length - len(Named_Insured_Mailing_Address)))
    Policy_Number.extend([None] * (max_length - len(Policy_Number)))
    Effective.extend([None] * (max_length - len(Effective)))

    # Create the DataFrame
    policy_data = pd.DataFrame({
        'Named Insured': Named_Insured_Mailing_Address,
        "Policy Number": Policy_Number,
        "Effective": Effective,
    }).drop_duplicates()
    print("Policy_data")
    print(policy_data)


    # Print the merged DataFrame to see the extracted information with the calculated differences
    print('merged_df')
    print(merged_df)
    excel_file = 'insurance_report_combined.xlsx'  
    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:  
        policy_data.to_excel(writer, index=False, sheet_name='Insurance Report', startrow=2)  
        
        # Access the workbook and the worksheet  
        workbook = writer.book  
        worksheet = writer.sheets['Insurance Report']  

        # Add headings for the policy report  
        report_heading = "Insurance Report"  
        worksheet.merge_cells('A1:C1')  
        cell = worksheet.cell(row=1, column=1)  
        cell.value = report_heading  
        cell.alignment = Alignment(horizontal='center', vertical='center')  
        cell.font = Font(bold=True)  

        # Add border style for the columns  
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))  
        
        min_width = 20  # Minimum column width  

        # Set column widths and apply styles  
        for col in worksheet.columns:  
            if not col:  # If the column is empty, skip  
                continue  
            
            max_length = 0  
            column_letter = get_column_letter(col[0].column)  

            for cell in col:  
                if cell.value:  # Check if the cell is not empty  
                    max_length = max(max_length, len(str(cell.value)))  # Update max_length  
                
            adjusted_width = max(max_length + 2, min_width)  # Add some extra space  
            worksheet.column_dimensions[column_letter].width = adjusted_width
            # Set borders for each cell in the column  
            for cell in col:  
                cell.border = border  

        # Make header row bold (now it’s row 3 after writing the policy data)  
        for cell in worksheet[2]:  # Row indexing starts from 1  
            cell.font = Font(bold=True)  # Make the header bold  

        # Loop to append additional DataFrames   
        last_row = worksheet.max_row + 2  
        for df in [merged_df, premium_df, coverage_df1, coverage_df2]:  
            df.to_excel(writer, index=False, sheet_name='Insurance Report', startrow=last_row) 
            
            

            # Update last_row after writing each DataFrame  
            last_row = worksheet.max_row + 2  

            # Apply the same styling and formatting to each appended DataFrame  
            for col in worksheet.iter_cols(min_row=last_row):  
                if not col:  # If the column is empty, skip  
                    continue  
                
                max_length = 0  
                column_letter = get_column_letter(col[0].column)
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    cell.border = border
        

                for cell in col:  
                    if cell.value:  # Only check non-empty cells  
                        max_length = max(max_length, len(str(cell.value)))  

                adjusted_width = max(max_length + 2, min_width)  
                worksheet.column_dimensions[column_letter].width = adjusted_width  

                # Apply borders  
                for cell in col:  
                    cell.border = border  

        # Save changes  
        workbook.save(excel_file)  

    print(f'Data successfully written to {excel_file}') 
 


      # Save the file with a unique name
    return excel_file

@app.route('/download1/<filename>', methods=['GET'])
def download_file1(filename):
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    return send_file(file_path, as_attachment=True)

@app.route('/view1/<filename>', methods=['GET'])
def view_file1(filename):
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    return send_file(file_path)

if __name__ == '__main__':
    os.makedirs('uploads', exist_ok=True)  # Create the uploads directory if it doesn't exist
    app.run(debug=True)

